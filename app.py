import streamlit as st
import pandas as pd
import numpy as np
import os
import io

# --- Helper Function ---
def parse_list(raw, is_multipliers=False):
    if not raw:
        return []
    items = [x.strip() for x in raw.replace("\n", " ").replace("\t", " ").replace(",", " ").split(" ") if x.strip()]
    if is_multipliers:
        items = [float(x) for x in items]
    return items

def normalize_geo(name: str):
    return str(name).strip().upper().replace(".", "").replace("_", "").replace(" ", "")

# --- 1. Master Recalculation Engine ---
def recalculate_state():
    """Rebuilds data strictly from the original copy, utilizing the MAP sheet for Geo matching."""
    ads = st.session_state.ads_original.copy()
    gran = {s: df.copy() for s, df in st.session_state.gran_original.items()} if st.session_state.gran_original else None
    
    # Build Map Dictionary
    map_dict = {}
    if st.session_state.map_df is not None:
        m_df = st.session_state.map_df
        if "MAP" in m_df.columns and "GEOGRAPHY" in m_df.columns:
            map_dict = dict(zip(
                m_df["MAP"].astype(str).str.strip().str.upper(), 
                m_df["GEOGRAPHY"].astype(str).str.strip().str.upper()
            ))

    for rule in st.session_state.rules:
        geo_val = rule["geography"]
        period_val = rule["period"]
        
        rule["ads_vars"] = []
        rule["gran_vars"] = []
        rule["missed"] = []
        rule["geo_warning"] = False
        
        # Determine valid Granular sheets based on MAP sheet logic
        valid_sheets = []
        if gran is not None:
            if geo_val == "All":
                valid_sheets = list(gran.keys())
            else:
                target_geo = normalize_geo(geo_val)
                for sheet_name in gran.keys():
                    sheet_upper = str(sheet_name).strip().upper()
                    mapped_geo = map_dict.get(sheet_upper, sheet_upper)
                    if normalize_geo(mapped_geo) == target_geo:
                        valid_sheets.append(sheet_name)
                
                if not valid_sheets:
                    rule["geo_warning"] = True

        for orig_var, mult in zip(rule["cols"], rule["mults"]):
            routed_to_granular = False
            
            # --- STRICT GRANULAR LOGIC ---
            if gran is not None and valid_sheets:
                gran_var = orig_var.strip().upper()
                target_period = str(period_val).strip().upper()

                for sheet_name in valid_sheets:
                    df = gran[sheet_name]
                    cols_upper = [str(c).strip().upper() for c in df.columns]
                    
                    if "VARIABLE" in cols_upper and "CONTRIBUTION" in cols_upper:
                        var_col = df.columns[cols_upper.index("VARIABLE")]
                        contrib_col = df.columns[cols_upper.index("CONTRIBUTION")]
                        
                        mask = (df[var_col].astype(str).str.strip().str.upper() == gran_var)
                        if period_val != "All":
                            mask &= (df[contrib_col].astype(str).str.strip().str.upper() == target_period)
                        
                        if mask.any():
                            min_col = df.columns[cols_upper.index("MIN")] if "MIN" in cols_upper else None
                            max_col = df.columns[cols_upper.index("MAX")] if "MAX" in cols_upper else None
                            
                            if min_col and max_col:
                                # FIX 1: Force float type globally to safely accept decimal changes
                                df[min_col] = pd.to_numeric(df[min_col], errors='coerce').astype(float)
                                df[max_col] = pd.to_numeric(df[max_col], errors='coerce').astype(float)
                                
                                df.loc[mask, min_col] = (df.loc[mask, min_col] * float(mult)).round(6)
                                df.loc[mask, max_col] = (df.loc[mask, max_col] * float(mult)).round(6)
                                
                                gran[sheet_name] = df
                                routed_to_granular = True
            
            # --- ADS LOGIC ---
            if routed_to_granular:
                rule["gran_vars"].append(f"{orig_var} (x{mult})")
            else:
                ads_var = orig_var if orig_var.endswith("_PMF") else orig_var + "_PMF"
                if ads_var in ads.columns:
                    # FIX 2: Cast entire column to float globally before partial row slice multiplication
                    ads[ads_var] = pd.to_numeric(ads[ads_var], errors='coerce').astype(float)
                    
                    ads_mask = pd.Series(True, index=ads.index)
                    if geo_val != "All":
                        ads_mask &= (ads[geo_col] == geo_val)
                    if period_val != "All":
                        ads_mask &= (ads[period_col] == period_val)

                    if ads_mask.any():
                        ads.loc[ads_mask, ads_var] = (ads.loc[ads_mask, ads_var] * float(mult)).round(6)
                        rule["ads_vars"].append(f"{ads_var} (x{mult})")
                    else:
                        rule["missed"].append(orig_var)
                else:
                    rule["missed"].append(orig_var)
                    
    st.session_state.ads = ads
    st.session_state.granular_sheets = gran

# --- Initialize Session State ---
if 'ads_original' not in st.session_state: st.session_state.ads_original = None
if 'gran_original' not in st.session_state: st.session_state.gran_original = None
if 'map_df' not in st.session_state: st.session_state.map_df = None
if 'ads' not in st.session_state: st.session_state.ads = None
if 'granular_sheets' not in st.session_state: st.session_state.granular_sheets = None
if 'rules' not in st.session_state: st.session_state.rules = []
if 'ads_last_uploaded' not in st.session_state: st.session_state.ads_last_uploaded = None
if 'gran_last_uploaded' not in st.session_state: st.session_state.gran_last_uploaded = None

st.set_page_config(page_title="Data Rule Modifier", layout="wide")
st.title("Data & Granular Rule Modifier")

# --- 2. File Upload ---
st.header("1. Upload Data")
col_up1, col_up2 = st.columns(2)

with col_up1:
    st.subheader("ADS Data")
    ads_file = st.file_uploader("Select ADS File (CSV/Excel)", type=["csv", "xlsx"], key="ads_upload")
    if ads_file is not None and st.session_state.ads_last_uploaded != ads_file.name:
        df = pd.read_csv(ads_file) if ads_file.name.endswith('.csv') else pd.read_excel(ads_file)
        st.session_state.ads_original = df.copy()
        st.session_state.ads = df.copy()
        st.session_state.ads_filename = os.path.splitext(ads_file.name)[0]
        st.session_state.ads_last_uploaded = ads_file.name
        st.session_state.rules = [] 
        st.success("✅ ADS loaded!")

with col_up2:
    st.subheader("Granular Spec")
    gran_file = st.file_uploader("Select Granular Spec (Excel)", type=["xlsx"], key="gran_upload")
    if gran_file is not None and st.session_state.gran_last_uploaded != gran_file.name:
        xl = pd.ExcelFile(gran_file)
        
        map_sheet_name = next((s for s in xl.sheet_names if "map" in s.lower()), None)
        if map_sheet_name:
            map_df = pd.read_excel(gran_file, sheet_name=map_sheet_name, dtype=str)
            map_df.columns = [str(c).strip().upper() for c in map_df.columns]
            st.session_state.map_df = map_df
        else:
            st.session_state.map_df = None
            
        gran_dict = {s: pd.read_excel(gran_file, sheet_name=s) for s in xl.sheet_names if s != map_sheet_name}
        
        st.session_state.gran_original = gran_dict
        st.session_state.granular_sheets = {s: df.copy() for s, df in gran_dict.items()}
        st.session_state.gran_filename = os.path.splitext(gran_file.name)[0]
        st.session_state.gran_last_uploaded = gran_file.name
        
        if map_sheet_name:
            st.success("✅ Granular Spec loaded (MAP Sheet Detected!)")
        else:
            st.warning("⚠️ Granular Spec loaded, but no 'MAP' sheet was found. Geography routing may rely on sheet names.")

# --- 3. Data Processing ---
if st.session_state.ads is not None:
    ads = st.session_state.ads
    
    st.divider()
    st.header("2. Apply Rules")

    default_geo = next((c for c in ["Geography","Geo","Region"] if c in ads.columns), None)
    default_period = next((c for c in ["Season","Time_Periods","Period_Definition"] if c in ads.columns), None)

    col_map1, col_map2 = st.columns(2)
    with col_map1:
        geo_col = st.selectbox("ADS Geography Column:", options=ads.columns, 
                               index=list(ads.columns).index(default_geo) if default_geo else 0)
    with col_map2:
        period_col = st.selectbox("ADS Period/Season Column:", options=ads.columns, 
                                  index=list(ads.columns).index(default_period) if default_period else 0)

    geo_options = ["All"] + list(st.session_state.ads_original[geo_col].dropna().unique())
    period_options = ["All"] + list(st.session_state.ads_original[period_col].dropna().unique())

    col1, col2 = st.columns(2)
    with col1:
        geo_val = st.selectbox("Geography:", options=geo_options)
        cols_text = st.text_input("Variables:", placeholder="e.g. Sales, Profit")
    with col2:
        period_val = st.selectbox("Time Period:", options=period_options)
        mults_text = st.text_input("Multipliers:", placeholder="e.g. 1.1, 0.9")

    # --- ACTION BUTTONS ---
    btn_col1, btn_col2, btn_col3 = st.columns([2, 1, 1])
    
    with btn_col1:
        if st.button("Apply Rule", type="primary", use_container_width=True):
            cols = parse_list(cols_text)
            try:
                mults = parse_list(mults_text, is_multipliers=True)
            except ValueError:
                st.error("⚠️ Error: Multipliers must be numbers.")
                st.stop()

            if not cols or not mults:
                st.warning("⚠️ Please enter both columns and multipliers.")
                st.stop()
            elif len(cols) != len(mults):
                st.error(f"⚠️ Mismatch: You have {len(cols)} columns but {len(mults)} multipliers.")
                st.stop()
            
            st.session_state.rules.append({
                "geography": geo_val,
                "period": period_val,
                "cols": cols,
                "mults": mults
            })
            recalculate_state()
            st.rerun()

    with btn_col2:
        if st.button("⏪ Undo Last Rule", use_container_width=True):
            if st.session_state.rules:
                st.session_state.rules.pop()
                recalculate_state()
                st.success("Last rule reverted!")
                st.rerun()
            else:
                st.warning("No rules to undo.")

    with btn_col3:
        if st.button("🔄 Reset All Data", type="secondary", use_container_width=True):
            st.session_state.rules = []
            recalculate_state()
            st.success("All data completely reset to original uploads!")
            st.rerun()

    # --- Show Dynamic Messages For The Latest Rule ---
    st.divider()
    if st.session_state.rules:
        latest = st.session_state.rules[-1]
        
        if latest.get("geo_warning"):
            st.warning(f"⚠️ **Geography '{latest['geography']}' was NOT found in the Granular MAP sheet.** Granular Spec was skipped.")
            
        if latest.get("gran_vars"):
            st.success(f"📂 **Applied to Granular Spec:** {', '.join(latest['gran_vars'])}")
        if latest.get("ads_vars"):
            st.success(f"📊 **Applied to ADS Dataset:** {', '.join(latest['ads_vars'])}")
        if latest.get("missed"):
            st.error(f"❌ **Failed to match anywhere (Skipped):** {', '.join(latest['missed'])}")

    # --- 4. Summary & Download ---
    st.divider()
    st.header("3. Summary & Download")
    
    if st.session_state.rules:
        st.subheader("Applied Rules Log")
        for i, r in enumerate(st.session_state.rules, start=1):
            with st.expander(f"Rule {i} | Geo: {r['geography']} | Period: {r['period']}", expanded=True):
                if r.get("geo_warning"):
                    st.markdown("⚠️ *Granular skipped: Geography missing from MAP sheet.*")
                if r.get('gran_vars'):
                    st.markdown(f"**📂 Sent to Granular Spec:** `{', '.join(r['gran_vars'])}`")
                if r.get('ads_vars'):
                    st.markdown(f"**📊 Sent to ADS:** `{', '.join(r['ads_vars'])}`")
                if r.get('missed'):
                    st.markdown(f"❌ **Failed to match:** `{', '.join(r['missed'])}`")
        st.divider()

    col_dl1, col_dl2 = st.columns(2)
    
    with col_dl1:
        st.subheader("ADS Download")
        ads_save_name = st.text_input("ADS Save As:", value=f"{st.session_state.ads_filename}_modified.csv")
        csv_data = st.session_state.ads.to_csv(index=False).encode('utf-8')
        st.download_button(label="Download Modified ADS (CSV)", data=csv_data, file_name=ads_save_name, mime='text/csv')

    with col_dl2:
        st.subheader("Granular Spec Download")
        if st.session_state.granular_sheets is not None:
            gran_save_name = st.text_input("Granular Save As:", value=f"{st.session_state.gran_filename}_modified.xlsx")
            
            output_buffer = io.BytesIO()
            with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
                if st.session_state.map_df is not None:
                    st.session_state.map_df.to_excel(writer, sheet_name="MAP", index=False)
                
                for sheet_name, df_sheet in st.session_state.granular_sheets.items():
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
                
                from openpyxl.styles import numbers
                for sheet_name in writer.book.sheetnames:
                    ws = writer.book[sheet_name]
                    if ws.max_row > 1:
                        header = [str(cell.value).strip().upper() if cell.value else "" for cell in ws[1]]
                        min_idx = header.index("MIN") + 1 if "MIN" in header else None
                        max_idx = header.index("MAX") + 1 if "MAX" in header else None
                        
                        if min_idx or max_idx:
                            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                                if min_idx and isinstance(row[min_idx - 1].value, (int, float)):
                                    row[min_idx - 1].number_format = numbers.FORMAT_PERCENTAGE_00
                                if max_idx and isinstance(row[max_idx - 1].value, (int, float)):
                                    row[max_idx - 1].number_format = numbers.FORMAT_PERCENTAGE_00
            
            output_bytes = output_buffer.getvalue()
            st.download_button(
                label="Download Modified Granular Spec (Excel)", 
                data=output_bytes, 
                file_name=gran_save_name, 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

else:
    st.info("⚠️ Please upload the ADS dataset to begin.")
